#!/usr/bin/env python3
# analytics_client_clean.py - Complete version with all export features

import requests
import json
from datetime import datetime
import sys
import argparse
import csv

class AnalyticsClient:
    def __init__(self, base_url="https://centralsoft.com.my"):
        self.base_url = base_url.rstrip('/')
        self.analytics_url = f"{self.base_url}/api/analytics"
        self.session = requests.Session()
        self.session.headers.update({
            'Content-Type': 'application/json',
            'User-Agent': 'CodeSquad-AnalyticsClient/1.0'
        })
        print(f"🔧 Analytics client initialized for: {self.base_url}")
    
    def test_connection(self):
        """Test if the backend is reachable."""
        print("🔍 Testing connection to backend...")
        try:
            response = self.session.get(
                f"{self.analytics_url}/dashboard/codesquad",
                timeout=5
            )
            if response.status_code == 200:
                print(f"✅ Connection successful! (Status: {response.status_code})")
                return True
            else:
                print(f"❌ Connection failed with status: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Connection failed: {e}")
            return False
    
    def get_dashboard(self, company_id):
        """Fetch all dashboard data for a specific company."""
        print(f"📡 Fetching dashboard data for: {company_id}")
        try:
            response = self.session.get(
                f"{self.analytics_url}/dashboard/{company_id}",
                timeout=10
            )
            if response.status_code == 200:
                data = response.json()
                print("✅ Dashboard data received successfully")
                return data
            else:
                print(f"❌ API returned error {response.status_code}")
                return None
        except Exception as e:
            print(f"❌ Request failed: {e}")
            return None
    
    def print_dashboard(self, company_id):
        """Print a formatted dashboard report."""
        print("\n" + "="*70)
        print("📊 CODE SQUAD ACCOUNTING DASHBOARD")
        print("="*70)
        
        dashboard_data = self.get_dashboard(company_id)
        
        if not dashboard_data or not dashboard_data.get('success'):
            print("❌ Failed to fetch dashboard data")
            return
        
        data = dashboard_data.get('data', {})
        company = data.get('company', {})
        summary = data.get('summary', {})
        recent_invoices = data.get('recent_invoices', [])
        
        # Company info
        print(f"\n🏢 COMPANY PROFILE")
        print(f"   Name: {company.get('companyName', 'N/A')}")
        print(f"   Industry: {company.get('industry', 'N/A')}")
        print(f"   LHDN TIN: {company.get('lhdnTinNo', 'N/A')}")
        
        # Summary - convert string numbers to float
        total_invoices = summary.get('total_invoices', 0) or 0
        
        total_revenue = summary.get('total_revenue', 0)
        if total_revenue is None:
            total_revenue = 0
        elif isinstance(total_revenue, str):
            total_revenue = float(total_revenue)
        
        avg_invoice = summary.get('avg_invoice', 0)
        if avg_invoice is None:
            avg_invoice = 0
        elif isinstance(avg_invoice, str):
            avg_invoice = float(avg_invoice)
        
        validated_count = summary.get('validated_count', 0) or 0
        
        print(f"\n📈 FINANCIAL SUMMARY")
        print(f"   Total Invoices: {total_invoices:,}")
        print(f"   Total Revenue: RM{total_revenue:,.2f}")
        print(f"   Average Invoice: RM{avg_invoice:,.2f}")
        print(f"   LHDN Validated: {validated_count}")
        
        # Recent invoices
        print(f"\n📋 RECENT INVOICES")
        if recent_invoices:
            for idx, invoice in enumerate(recent_invoices[:5], 1):
                invoice_no = invoice.get('invoice_no', 'N/A')
                partner = invoice.get('partner_name', 'N/A')
                
                amount = invoice.get('net_amount', 0)
                if amount is None:
                    amount = 0
                elif isinstance(amount, str):
                    amount = float(amount)
                
                status = invoice.get('lhdn_status', 'N/A')
                
                if len(partner) > 20:
                    partner = partner[:17] + "..."
                
                print(f"   {idx}. {invoice_no} | {partner} | RM{amount:,.2f} | {status}")
        else:
            print("   No recent invoices found")
        
        print(f"\n🕐 Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*70 + "\n")
    
    def save_report_to_file(self, company_id, filename=None):
        """Save dashboard data to JSON file."""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"dashboard_{company_id}_{timestamp}.json"
        
        dashboard_data = self.get_dashboard(company_id)
        if dashboard_data:
            try:
                with open(filename, 'w') as f:
                    json.dump(dashboard_data, f, indent=2)
                print(f"💾 JSON report saved to: {filename}")
                return True
            except Exception as e:
                print(f"❌ Failed to save: {e}")
                return False
        return False
    
    def export_to_csv(self, company_id, filename=None):
        """Export recent invoices to CSV file."""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"invoices_{company_id}_{timestamp}.csv"
        
        dashboard_data = self.get_dashboard(company_id)
        if not dashboard_data or not dashboard_data.get('success'):
            print("❌ No data to export")
            return False
        
        invoices = dashboard_data['data'].get('recent_invoices', [])
        if not invoices:
            print("❌ No invoices found to export")
            return False
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                # Get all column names from the first invoice
                fieldnames = list(invoices[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Write header
                writer.writeheader()
                
                # Write data rows
                writer.writerows(invoices)
            
            print(f"💾 Exported {len(invoices)} invoices to: {filename}")
            return True
        except Exception as e:
            print(f"❌ Failed to export CSV: {e}")
            return False
    
    def export_to_excel(self, company_id, filename=None):
        """Export dashboard data to Excel file using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("❌ openpyxl not installed. Run: pip3 install openpyxl")
            return False
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"dashboard_{company_id}_{timestamp}.xlsx"
        
        dashboard_data = self.get_dashboard(company_id)
        if not dashboard_data or not dashboard_data.get('success'):
            print("❌ No data to export")
            return False
        
        data = dashboard_data['data']
        company = data.get('company', {})
        summary = data.get('summary', {})
        invoices = data.get('recent_invoices', [])
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add company info
        ws_summary['A1'] = "CODE SQUAD ACCOUNTING DASHBOARD"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A3'] = "Company Name:"
        ws_summary['B3'] = company.get('companyName', 'N/A')
        ws_summary['A4'] = "Industry:"
        ws_summary['B4'] = company.get('industry', 'N/A')
        ws_summary['A5'] = "LHDN TIN:"
        ws_summary['B5'] = company.get('lhdnTinNo', 'N/A')
        
        # Add summary stats
        ws_summary['A7'] = "FINANCIAL SUMMARY"
        ws_summary['A7'].font = Font(bold=True)
        ws_summary['A8'] = "Total Invoices:"
        ws_summary['B8'] = summary.get('total_invoices', 0)
        ws_summary['A9'] = "Total Revenue:"
        total_revenue = summary.get('total_revenue', 0)
        if isinstance(total_revenue, str):
            total_revenue = float(total_revenue)
        ws_summary['B9'] = total_revenue
        ws_summary['B9'].number_format = 'RM #,##0.00'
        ws_summary['A10'] = "Average Invoice:"
        avg_invoice = summary.get('avg_invoice', 0)
        if isinstance(avg_invoice, str):
            avg_invoice = float(avg_invoice)
        ws_summary['B10'] = avg_invoice
        ws_summary['B10'].number_format = 'RM #,##0.00'
        ws_summary['A11'] = "LHDN Validated:"
        ws_summary['B11'] = summary.get('validated_count', 0)
        
        # Sheet 2: Invoices
        ws_invoices = wb.create_sheet("Invoices")
        
        if invoices:
            # Add headers
            headers = list(invoices[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_invoices.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row, invoice in enumerate(invoices, 2):
                for col, header in enumerate(headers, 1):
                    value = invoice.get(header, '')
                    if header == 'net_amount' and value:
                        try:
                            value = float(value) if isinstance(value, str) else value
                        except:
                            pass
                    ws_invoices.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for col in ws_invoices.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_invoices.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        try:
            wb.save(filename)
            print(f"💾 Excel report saved to: {filename}")
            return True
        except Exception as e:
            print(f"❌ Failed to save Excel file: {e}")
            return False
    
    def export_to_excel_pandas(self, company_id, filename=None):
        """Export to Excel using pandas (more reliable)."""
        try:
            import pandas as pd
        except ImportError:
            print("❌ pandas not installed. Run: pip3 install pandas")
            return False
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"dashboard_{company_id}_{timestamp}.xlsx"
        
        dashboard_data = self.get_dashboard(company_id)
        if not dashboard_data or not dashboard_data.get('success'):
            print("❌ No data to export")
            return False
        
        data = dashboard_data['data']
        invoices = data.get('recent_invoices', [])
        
        # Convert to DataFrame
        df = pd.DataFrame(invoices)
        
        # Convert amount to float
        if 'net_amount' in df.columns:
            df['net_amount'] = pd.to_numeric(df['net_amount'], errors='coerce')
        
        # Create Excel file with multiple sheets
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Invoices sheet
                df.to_excel(writer, sheet_name='Invoices', index=False)
                
                # Summary sheet
                summary_data = {
                    'Metric': ['Total Invoices', 'Total Revenue', 'Average Invoice', 'LHDN Validated'],
                    'Value': [
                        data['summary'].get('total_invoices', 0),
                        data['summary'].get('total_revenue', 0),
                        data['summary'].get('avg_invoice', 0),
                        data['summary'].get('validated_count', 0)
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                
                # Statistics sheet
                if not df.empty and 'net_amount' in df.columns:
                    stats_data = {
                        'Statistic': ['Highest Invoice', 'Lowest Invoice', 'Average Invoice', 'Total Revenue'],
                        'Amount': [
                            f"RM {df['net_amount'].max():,.2f}",
                            f"RM {df['net_amount'].min():,.2f}",
                            f"RM {df['net_amount'].mean():,.2f}",
                            f"RM {df['net_amount'].sum():,.2f}"
                        ]
                    }
                    pd.DataFrame(stats_data).to_excel(writer, sheet_name='Statistics', index=False)
            
            print(f"💾 Excel report (pandas) saved to: {filename}")
            return True
        except Exception as e:
            print(f"❌ Failed to save Excel file: {e}")
            return False


def main():
    print("\n🚀 Code Squad Analytics Client")
    print("="*40)
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Code Squad Analytics Client')
    parser.add_argument('company', nargs='?', default='codesquad', help='Company ID')
    parser.add_argument('--url', default='https://centralsoft.com.my', help='Backend URL')
    parser.add_argument('--save', action='store_true', help='Save JSON report')
    parser.add_argument('--csv', action='store_true', help='Export to CSV')
    parser.add_argument('--excel', action='store_true', help='Export to Excel (openpyxl)')
    parser.add_argument('--excel-pandas', action='store_true', help='Export to Excel (pandas)')
    
    args = parser.parse_args()
    
    client = AnalyticsClient(args.url)
    
    if not client.test_connection():
        print("\n💡 Troubleshooting: Check if backend is running")
        return
    
    # Display dashboard
    client.print_dashboard(args.company)
    
    # Export based on flags
    if args.save:
        client.save_report_to_file(args.company)
    
    if args.csv:
        client.export_to_csv(args.company)
    
    if args.excel:
        client.export_to_excel(args.company)
    
    if args.excel_pandas:
        client.export_to_excel_pandas(args.company)


if __name__ == "__main__":
    main()